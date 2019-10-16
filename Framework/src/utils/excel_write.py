# -*- coding: utf-8 -*-
# @Author: yusi
# @Date:   2019-07-17 10:16:49
# @Last Modified by:   yusi
# @Last Modified time: 2019-07-19 10:28:31
"""写入excel"""
import openpyxl as xl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, colors, Side, Border
from openpyxl.utils import get_column_letter, column_index_from_string
import re,xlwt,os

def copy_excel(excelpath1, excelpath2):
    '''复制excek，把excelpath1数据复制到excelpath2'''
    wb2 = xl.Workbook()
    wb2.save(excelpath2)
    # 读取数据
    wb1 = xl.load_workbook(excelpath1)
    wb2 = xl.load_workbook(excelpath2)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]
    max_row = sheet1.max_row         # 最大行数
    max_column = sheet1.max_column   # 最大列数

    for m in list(range(1,max_row+1)):
        for n in list(range(97,97+max_column)):   # chr(97)='a'
            n = chr(n)                            # ASCII字符
            i ='%s%d'% (n, m)                     # 单元格编号
            cell1 = sheet1[i].value               # 获取data单元格数据
            sheet2[i].value = cell1               # 赋值到test单元格

    wb2.save(excelpath2)                 # 保存数据
    wb1.close()                          # 关闭excel
    wb2.close()



def writeExcel(path, data, title=None, row_height=18, merge_row_col=[]):
    """
    path：写入excel的文件路径，不存在新建，存在的话就覆盖

    data：写入excel的数据，格式为[[], [],...]或([], [],...)或[(), (),...]或(()，( ),...)，第二层每个元素表示一行数据，可以长度不一致，表格的列数以最长的那个元素为准，值为None或者长度达不到最达长度的，写入excel时为空

    title：表格名称，就是我们excel中的sheet1,sheet2,类型为字符串

    row_height：设置表格行高，默认为18，

    merge_row_col：合并单元格的起止位置组合，格式为({},{},...)或[{},{},...]，其中元素{}的key-value
    """
    if not isinstance(data, list) and not isinstance(data, tuple):
        raise BaseException("data只能是列表或元组")
    if len(data) > 0 and (not isinstance(data[0], list) and not isinstance(data[0], tuple)):
        raise BaseException("data只能是列表或元组")
    if len(merge_row_col) > 0:
        if not isinstance(merge_row_col, tuple) and not isinstance(merge_row_col, list):
            raise BaseException("merge_row_col只能是以dict为元素的列表或元组")
        elif not isinstance(merge_row_col[0], dict):
            raise BaseException("merge_row_col只能是以dict为元素的列表或元组")
        if not judge_merge(merge_row_col):
            raise BaseException("merge_row_col不符合要求，不能合并")
    workbook = xl.Workbook()
    sheet = workbook.active
    if title:
        sheet = workbook.create_sheet(title=title, index=0)
    maxRow = len(data)
    maxCol = 0
    for row in data:
        maxCol = len(row) if maxCol < len(row) else maxCol
        sheet.append(row)
    col_length = {
 
    }

    
    '''
    合并单元格,放在设置单元格样式之前，否则无边框
    '''
    for row_col in merge_row_col:
        sheet.merge_cells(
            start_row=row_col["start_row"],
            start_column=row_col["start_column"],
            end_row=row_col["end_row"],
            end_column=row_col["end_column"]
        )
 
    '''
    设置单元格样式
    '''
    for r in range(1, maxRow + 1):
        sheet.row_dimensions[r].height = row_height
        for c in range(1, maxCol + 1):
            cell = sheet.cell(row=r, column=c)
            # 设置边框
            cell.border = getBorder()
            re_hanzi = re.compile(r"[\u4E00-\u9FFF]")
            w = len(str(cell.value))
            # 有汉字乘以2，否则乘以1.5
            if re_hanzi.search(str(cell.value)):
                w *= 2
            else:
                w *= 1.5
            # 列宽最小为8，最大为38，超过38文字缩小至适合大小
            w = 8 if w < 8 else w
            if w > 38:
                w = 38
                cell.alignment = getAlignment(shrink_to_fit=True)
            else:
                cell.alignment = getAlignment()
            if col_length.get(cell.column):
                col_length[cell.column] = w if col_length[cell.column] < w else col_length[cell.column]
            else:
                col_length[cell.column] = w
            # 第一行加粗
            if r == 1:
                cell.font = getFont(bold=True)
            else:
                cell.font = getFont()
    '''
    设置列宽
    '''
    for key, val in col_length.items():
        i =get_column_letter(key)
        sheet.column_dimensions[i].width = val + 0.7
 
    workbook.save(path)
 
 
"""
判断合并单元格是否符合要求
"""
 
 
def judge_merge(merge_row_col):
    if len(merge_row_col) == 1:
        if merge_row_col[0]["start_row"] <= merge_row_col[0]["end_row"] and \
                merge_row_col[0]["start_column"] <= merge_row_col[0]["end_column"]:
            return True
        else:
            return False
    for i in range(len(merge_row_col)):
        tempi = merge_row_col[i]
        if not (tempi["start_row"] <= tempi["end_row"] and tempi["start_column"] <= tempi["end_column"]):
            return False
        if i == len(merge_row_col) - 1:
            break
        for j in range(i + 1, len(merge_row_col)):
            tempj = merge_row_col[j]
            if not (tempj["start_column"] > tempi["end_column"] or \
                    tempj["end_column"] < tempi["start_column"] or \
                    tempj["start_row"] > tempi["end_row"] or \
                    tempj["end_row"] < tempi["start_row"]):
                return False
    return True
 
 
def getAlignment(horizontal='center', vertical='center', shrink_to_fit=False, wrap_text=False, text_rotation=0,
                 indent=0):
    agt = Alignment(
        horizontal=horizontal,
        vertical=vertical,
        shrink_to_fit=shrink_to_fit,
        wrap_text=wrap_text,
        text_rotation=text_rotation,
        indent=indent
    )
    return agt
 
 
def getFont(name='微软雅黑', size=12, bold=False, italic=False, vertAlign=None, underline=None, strike=False,
            color=colors.BLACK):
    font = Font(
        name=name,
        size=size,
        bold=bold,
        italic=italic,
        vertAlign=vertAlign,
        underline=underline,
        strike=strike,
        color=color
    )
    return font
 
 
'''
style的值为：
{'double', 'dashDot', 'mediumDashDot', 'dotted', 'medium', 'thick', 'dashDotDot', 'thin', 
'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'dashed', 'hair'}
'''
 
 
def getBorder():
    border = Border(
        left=Side(style='thin', color='FF000000'),
        right=Side(style='thin', color='FF000000'),
        top=Side(style='thin', color='FF000000'),
        bottom=Side(style='thin', color='FF000000')
    )
    return border

if __name__ == "__main__":
    data = (
        ["ID", "name", "age", "createDate", "title1", "title2"],
        [1, "杨", 23, "2018-12-09","rewf0"],
        [2, "张", 33, "2018-12-09", None, "很长很长的中文文字很长很长的中文文字很长很长的中文文字很长很长的中文文字",],
        [3, "王", 42, "2018-12-10","长中文文字长中文文字长中文文字长中文文字"]
    )
    merge_row_col = (
        {
            "start_row": 2,
            "start_column": 4,
            "end_row": 3,
            "end_column": 4
        },
    )
    # BASE_PATH = os.path.split(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))[0]
    # REPORT_PATH = os.path.join(BASE_PATH, 'report')
    # path = REPORT_PATH +'\\test.xlsx'
    # writeExcel(path=path, data=data, merge_row_col=merge_row_col)
